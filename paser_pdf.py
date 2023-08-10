import openpyxl as xl
import pdfplumber
import os

# Buscar os arquivos pdf:
for arquivo in os.listdir('pdfs'):
    
    if arquivo.lower().endswith('.pdf'):
        try:
            excel = xl.load_workbook('Base de dados inspecoes.xlsx')
            aba = excel.active
            linha_inicio = len(aba['A']) + 1

            pdf = pdfplumber.open(f'pdfs\\{arquivo}')
            pagina =pdf.pages[0]
            dados = pagina.extract_table()

            for indice, dado in enumerate(dados[1:], start = linha_inicio):

                if dado[0]=="":
                    pass
                else:
                    aba.cell(row=indice, column=1).value = dado[0]
                    aba.cell(row=indice, column=2).value = dado[1]
                    aba.cell(row=indice, column=3).value = dado[2]
                    aba.cell(row=indice, column=4).value = dado[3]
                    aba.cell(row=indice, column=5).value = dado[4]
                

            pdf.close()
            excel.save('Base de dados inspecoes.xlsx')
            excel.close()
        except:
            with open('log_erros.txt', 'a') as log:
                log.write(f'Ocorreu um erro ao tentar abrir o arquivo: {arquivo}. Ele não foi adicionado a planilha Excel.\n')
    else:
        with open('log_erros.txt', 'a') as log:
            log.write(f'O arquivo: {arquivo} não foi adicionado a planilha Excel pois não é um arquivo PDF válido.\n')